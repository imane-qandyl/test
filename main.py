#!/usr/bin/env python3
from pathlib import Path
import argparse
import time
import pandas as pd

from src.config import CATEGORY_DICTIONARY
from src.classifiers import recategorize_dataset
from src.extractors import extract_dimensions_series, extract_colors_series


def quick_format_columns(output_file: Path) -> None:
    """Format all column widths based on entire dataset for professional appearance."""
    from openpyxl import load_workbook
    
    workbook = load_workbook(output_file)
    sheet = workbook.active

    for column in sheet.columns:
        max_len = len(str(column[0].value or ''))
        
        # Check all rows in column for accurate width
        for cell in column[1:]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        
        col_letter = column[0].column_letter
        width = min(max(12, max_len + 2), 50)
        sheet.column_dimensions[col_letter].width = width

    workbook.save(output_file)


def process_file(input_path: Path, output_dir: Path) -> None:
    input_path = Path(input_path)
    output_dir = Path(output_dir)
    if not input_path.exists():
        raise FileNotFoundError(f"Source file not found: {input_path}")

    start_total = time.time()

    print("📥 Chargement du dataset...")
    start = time.time()
    df = pd.read_excel(input_path, engine='pyxlsb')
    print(f"   ⏱️  {time.time() - start:.2f}s")

    print("🔄 Recherche des anomalies de catégories...")
    start = time.time()
    df = recategorize_dataset(df, CATEGORY_DICTIONARY)
    print(f"   ⏱️  {time.time() - start:.2f}s")

    print("🎨 Extraction des dimensions et des couleurs en cours...")
    start = time.time()
    df['Extracted_Dimension'] = extract_dimensions_series(df['Libellé produit'])
    df['Extracted_Color'] = extract_colors_series(df['Libellé produit'])
    print(f"   ⏱️  {time.time() - start:.2f}s")

    if 'Cod_cmd' in df.columns:
        df['Cod_cmd'] = df['Cod_cmd'].astype(str)

    if 'Date de commande' in df.columns:
        df['Date de commande'] = pd.to_datetime(
            df['Date de commande'], unit='D', origin='1899-12-30', errors='coerce'
        ).dt.strftime('%d/%m/%Y')

    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"{input_path.stem} CLEAN.xlsx"

    print("💾 Enregistrement du fichier de données nettoyé...")
    start = time.time()
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Données')
        ws = writer.sheets['Données']
        ws.freeze_panes = 'A2'
        for i, col in enumerate(df.columns, start=1):
            max_len = max(int(df[col].astype(str).str.len().max()) if len(df) else 0, len(str(col)))
            width = min(max(max_len + 2, 12), 50)
            ws.column_dimensions[get_column_letter(i)].width = width
    print(f"   ⏱️  {time.time() - start:.2f}s")

def main() -> None:
    parser = argparse.ArgumentParser(description="Clean ecommerce dataset")
    parser.add_argument(
        "input",
        nargs='?',
        default="data/raw/20210614 Ecommerce sales.xlsb",
        help="Input .xlsb file path",
    )
    parser.add_argument("--out", "-o", default="data/processed", help="Output directory")
    args = parser.parse_args()

    process_file(Path(args.input), Path(args.out))


if __name__ == "__main__":
    main()